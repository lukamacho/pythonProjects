import openpyxl as opener
from openpyxl.chart import BarChart,Reference


def process_workbook(filename):

    workbook = opener.load_workbook(filename)
    sheet = workbook['Sheet1']

    for row in range(2,sheet.max_row+1):
        cell = sheet.cell(row,3)
        changed_price = cell.value * 0.9
        new_cell = sheet.cell(row,4)
        new_cell.value = changed_price

    values = Reference(sheet,min_row=2,max_row=4,min_col=4,max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')
    workbook.save('transactions2.xlsx')

process_workbook('transactions.xlsx')